import streamlit as st
import pandas as pd
import tempfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ⚠️ Render 환경에서는 상대 import 기준
from income import get_income_by_name  # income.py가 같은 폴더에 있어야 함

st.set_page_config(page_title="DART 재무제표 추출기", layout="wide")
st.title("📊 DART 재무제표 추출기")

# 입력창
corp_name = st.text_input("기업명", "삼성전자")
corp_market = st.text_input("기업 구분 ('Y': 코스피, 'K': 코스닥, 'N': 코넥스, 'E': 기타)", "Y")
bgn_de = st.text_input("시작일 (YYYYMMDD)", "20220101")
end_de = st.text_input("종료일 (YYYYMMDD)", "20241231")

# 버튼
if st.button("📥 엑셀 파일 생성 및 다운로드"):
    with st.spinner("📡 데이터를 조회하고 있습니다..."):
        try:
            # 임시 엑셀 파일 생성
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                filepath = tmp.name

            # income.py의 get_income_by_name 함수가 엑셀 파일 생성을 직접 담당
            generated_filepath = get_income_by_name(
                corp_name=corp_name,
                corp_market=corp_market,
                bgn_de=bgn_de,
                end_de=end_de,
                filepath=filepath  # 임시 파일 경로 전달
            )

            # 엑셀 파일 서식 적용
            wb = load_workbook(generated_filepath)
            
            # 셀 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid") # 진한 녹색
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            value_alignment = Alignment(horizontal="right", vertical="center")
            
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                
                # 헤더 스타일 적용
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # label_ko 열 (A열) 너비 설정
                ws.column_dimensions['A'].width = 40

                # 나머지 데이터 열에 대한 서식 및 너비 조정
                for col_idx in range(2, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    max_width = 10  # 최소 너비 설정

                    # 헤더 너비 고려
                    header_cell_value = str(ws.cell(row=1, column=col_idx).value)
                    max_width = max(max_width, len(header_cell_value) + 2) # 헤더 텍스트 길이 + 여백

                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'  # 천 단위 구분 기호
                            formatted_value = f"{cell.value:,.0f}"
                            max_width = max(max_width, len(formatted_value) + 2)
                            cell.alignment = value_alignment
                        elif cell.value is not None:
                            # 숫자 외 다른 텍스트도 너비 고려
                            max_width = max(max_width, len(str(cell.value)) + 2)
                        
                        cell.border = thin_border # 모든 데이터 셀에 테두리 적용

                    ws.column_dimensions[col_letter].width = max_width

            wb.save(generated_filepath)

            # 다운로드 버튼
            with open(generated_filepath, 'rb') as f:
                st.success("✅ 엑셀 파일이 준비되었습니다!")
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=f,
                    file_name=f"{corp_name}_재무제표.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"❌ 오류 발생: {e}")